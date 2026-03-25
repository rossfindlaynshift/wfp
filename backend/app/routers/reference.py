from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from openpyxl import load_workbook

from app.database import get_db
from app.models.reference import (
    FxRate,
    LoadingMultiplier,
    DepartmentHierarchy,
    EmployeeExclusion,
    MeritAssumption,
    BudgetTarget,
    QuotaRampSchedule,
    LeaveCostRate,
)
from app.schemas.reference import (
    FxRateCreate, FxRateOut,
    LoadingMultiplierCreate, LoadingMultiplierOut,
    DepartmentHierarchyCreate, DepartmentHierarchyOut,
    EmployeeExclusionCreate, EmployeeExclusionOut, EmployeeExclusionUpdate,
    MeritAssumptionCreate, MeritAssumptionOut,
    BudgetTargetCreate, BudgetTargetOut,
    QuotaRampScheduleCreate, QuotaRampScheduleOut,
    LeaveCostRateCreate, LeaveCostRateOut,
)

router = APIRouter()


# =============================================================================
# Generic CRUD helpers
# =============================================================================
def _get_all(db: Session, model):
    return db.query(model).all()


def _create(db: Session, model, data):
    obj = model(**data.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


def _update(db: Session, model, item_id: int, data):
    obj = db.get(model, item_id)
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(obj, key, value)
    db.commit()
    db.refresh(obj)
    return obj


def _delete(db: Session, model, item_id: int):
    obj = db.get(model, item_id)
    if not obj:
        raise HTTPException(status_code=404, detail="Not found")
    db.delete(obj)
    db.commit()
    return {"deleted": True}


def _bulk_upsert(db: Session, model, items: list, unique_key: str):
    """Replace all rows: delete existing, insert new."""
    db.query(model).delete()
    for item in items:
        obj = model(**item.model_dump())
        db.add(obj)
    db.commit()
    return db.query(model).all()


# =============================================================================
# FX Rates
# =============================================================================
@router.get("/fx-rates", response_model=list[FxRateOut])
def list_fx_rates(db: Session = Depends(get_db)):
    return _get_all(db, FxRate)


@router.post("/fx-rates", response_model=FxRateOut)
def create_fx_rate(data: FxRateCreate, db: Session = Depends(get_db)):
    return _create(db, FxRate, data)


@router.put("/fx-rates/{item_id}", response_model=FxRateOut)
def update_fx_rate(item_id: int, data: FxRateCreate, db: Session = Depends(get_db)):
    return _update(db, FxRate, item_id, data)


@router.delete("/fx-rates/{item_id}")
def delete_fx_rate(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, FxRate, item_id)


@router.put("/fx-rates", response_model=list[FxRateOut])
def bulk_replace_fx_rates(data: list[FxRateCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, FxRate, data, "currency")


# =============================================================================
# Loading Multipliers
# =============================================================================
@router.get("/loading-multipliers", response_model=list[LoadingMultiplierOut])
def list_loading_multipliers(db: Session = Depends(get_db)):
    return _get_all(db, LoadingMultiplier)


@router.post("/loading-multipliers", response_model=LoadingMultiplierOut)
def create_loading_multiplier(data: LoadingMultiplierCreate, db: Session = Depends(get_db)):
    return _create(db, LoadingMultiplier, data)


@router.put("/loading-multipliers/{item_id}", response_model=LoadingMultiplierOut)
def update_loading_multiplier(item_id: int, data: LoadingMultiplierCreate, db: Session = Depends(get_db)):
    return _update(db, LoadingMultiplier, item_id, data)


@router.delete("/loading-multipliers/{item_id}")
def delete_loading_multiplier(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, LoadingMultiplier, item_id)


@router.put("/loading-multipliers", response_model=list[LoadingMultiplierOut])
def bulk_replace_loading_multipliers(data: list[LoadingMultiplierCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, LoadingMultiplier, data, "legal_entity")


# =============================================================================
# Department Hierarchy
# =============================================================================
@router.get("/department-hierarchy", response_model=list[DepartmentHierarchyOut])
def list_department_hierarchy(db: Session = Depends(get_db)):
    return _get_all(db, DepartmentHierarchy)


@router.post("/department-hierarchy", response_model=DepartmentHierarchyOut)
def create_department_hierarchy(data: DepartmentHierarchyCreate, db: Session = Depends(get_db)):
    return _create(db, DepartmentHierarchy, data)


@router.put("/department-hierarchy/{item_id}", response_model=DepartmentHierarchyOut)
def update_department_hierarchy(item_id: int, data: DepartmentHierarchyCreate, db: Session = Depends(get_db)):
    return _update(db, DepartmentHierarchy, item_id, data)


@router.delete("/department-hierarchy/{item_id}")
def delete_department_hierarchy(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, DepartmentHierarchy, item_id)


@router.put("/department-hierarchy", response_model=list[DepartmentHierarchyOut])
def bulk_replace_department_hierarchy(data: list[DepartmentHierarchyCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, DepartmentHierarchy, data, "department")


# =============================================================================
# Employee Exclusions
# =============================================================================
@router.get("/employee-exclusions", response_model=list[EmployeeExclusionOut])
def list_employee_exclusions(db: Session = Depends(get_db)):
    return _get_all(db, EmployeeExclusion)


@router.post("/employee-exclusions", response_model=EmployeeExclusionOut)
def create_employee_exclusion(data: EmployeeExclusionCreate, db: Session = Depends(get_db)):
    return _create(db, EmployeeExclusion, data)


@router.put("/employee-exclusions/{item_id}", response_model=EmployeeExclusionOut)
def update_employee_exclusion(item_id: int, data: EmployeeExclusionUpdate, db: Session = Depends(get_db)):
    return _update(db, EmployeeExclusion, item_id, data)


@router.delete("/employee-exclusions/{item_id}")
def delete_employee_exclusion(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, EmployeeExclusion, item_id)


# =============================================================================
# Merit Assumptions
# =============================================================================
@router.get("/merit-assumptions", response_model=list[MeritAssumptionOut])
def list_merit_assumptions(db: Session = Depends(get_db)):
    return _get_all(db, MeritAssumption)


@router.post("/merit-assumptions", response_model=MeritAssumptionOut)
def create_merit_assumption(data: MeritAssumptionCreate, db: Session = Depends(get_db)):
    return _create(db, MeritAssumption, data)


@router.put("/merit-assumptions/{item_id}", response_model=MeritAssumptionOut)
def update_merit_assumption(item_id: int, data: MeritAssumptionCreate, db: Session = Depends(get_db)):
    return _update(db, MeritAssumption, item_id, data)


@router.delete("/merit-assumptions/{item_id}")
def delete_merit_assumption(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, MeritAssumption, item_id)


# =============================================================================
# Budget Targets
# =============================================================================
@router.get("/budget-targets", response_model=list[BudgetTargetOut])
def list_budget_targets(db: Session = Depends(get_db)):
    return _get_all(db, BudgetTarget)


@router.post("/budget-targets", response_model=BudgetTargetOut)
def create_budget_target(data: BudgetTargetCreate, db: Session = Depends(get_db)):
    return _create(db, BudgetTarget, data)


@router.put("/budget-targets/{item_id}", response_model=BudgetTargetOut)
def update_budget_target(item_id: int, data: BudgetTargetCreate, db: Session = Depends(get_db)):
    return _update(db, BudgetTarget, item_id, data)


@router.delete("/budget-targets/{item_id}")
def delete_budget_target(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, BudgetTarget, item_id)


@router.put("/budget-targets", response_model=list[BudgetTargetOut])
def bulk_replace_budget_targets(data: list[BudgetTargetCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, BudgetTarget, data, "l2_function")


# =============================================================================
# Quota Ramp Schedule
# =============================================================================
@router.get("/quota-ramp-schedule", response_model=list[QuotaRampScheduleOut])
def list_quota_ramp_schedule(db: Session = Depends(get_db)):
    return db.query(QuotaRampSchedule).order_by(QuotaRampSchedule.day_from).all()


@router.put("/quota-ramp-schedule", response_model=list[QuotaRampScheduleOut])
def replace_quota_ramp_schedule(data: list[QuotaRampScheduleCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, QuotaRampSchedule, data, "day_from")


# =============================================================================
# Leave Cost Rates
# =============================================================================
@router.get("/leave-cost-rates", response_model=list[LeaveCostRateOut])
def list_leave_cost_rates(db: Session = Depends(get_db)):
    return _get_all(db, LeaveCostRate)


@router.post("/leave-cost-rates", response_model=LeaveCostRateOut)
def create_leave_cost_rate(data: LeaveCostRateCreate, db: Session = Depends(get_db)):
    return _create(db, LeaveCostRate, data)


@router.put("/leave-cost-rates/{item_id}", response_model=LeaveCostRateOut)
def update_leave_cost_rate(item_id: int, data: LeaveCostRateCreate, db: Session = Depends(get_db)):
    return _update(db, LeaveCostRate, item_id, data)


@router.delete("/leave-cost-rates/{item_id}")
def delete_leave_cost_rate(item_id: int, db: Session = Depends(get_db)):
    return _delete(db, LeaveCostRate, item_id)


@router.put("/leave-cost-rates", response_model=list[LeaveCostRateOut])
def bulk_replace_leave_cost_rates(data: list[LeaveCostRateCreate], db: Session = Depends(get_db)):
    return _bulk_upsert(db, LeaveCostRate, data, "country")


# =============================================================================
# Excel Upload (generic for any ref table)
# =============================================================================
TABLE_CONFIGS = {
    "fx-rates": {
        "model": FxRate,
        "columns": {"currency": str, "rate_to_eur": float},
    },
    "loading-multipliers": {
        "model": LoadingMultiplier,
        "columns": {
            "legal_entity": str, "default_location": str,
            "employer_benefits_pct": float, "employer_taxes_pct": float,
            "currency": str,
        },
    },
    "department-hierarchy": {
        "model": DepartmentHierarchy,
        "columns": {"department": str, "l1": str, "l2": str, "l3": str},
    },
    "employee-exclusions": {
        "model": EmployeeExclusion,
        "columns": {
            "hibob_id": int, "display_name": str,
            "job_title": str, "department": str, "excluded": bool,
        },
    },
    "leave-cost-rates": {
        "model": LeaveCostRate,
        "columns": {"country": str, "cost_pct": float},
    },
    "budget-targets": {
        "model": BudgetTarget,
        "columns": {
            "l2_function": str, "month": str,
            "headcount": float, "loaded_cost": float, "pnl_cost": float,
        },
    },
}


@router.post("/upload/{table_name}")
async def upload_ref_table(
    table_name: str,
    file: UploadFile = File(...),
    replace: bool = Query(default=True, description="Replace all existing rows"),
    db: Session = Depends(get_db),
):
    if table_name not in TABLE_CONFIGS:
        raise HTTPException(status_code=400, detail=f"Unknown table: {table_name}. Valid: {list(TABLE_CONFIGS.keys())}")

    config = TABLE_CONFIGS[table_name]
    model = config["model"]
    expected_columns = config["columns"]

    contents = await file.read()
    wb = load_workbook(BytesIO(contents), data_only=True)
    ws = wb.active

    # Read header row
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    headers = [h.strip().lower().replace(" ", "_") if h else "" for h in headers]

    # Map expected columns to their positions
    col_map = {}
    for col_name in expected_columns:
        # Try exact match first, then fuzzy
        normalized = col_name.lower().replace("_", " ")
        for i, h in enumerate(headers):
            if h == col_name or h.replace("_", " ") == normalized:
                col_map[col_name] = i
                break

    missing = set(expected_columns) - set(col_map)
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing columns: {missing}. Found headers: {headers}",
        )

    # Parse rows
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data = {}
        skip = False
        for col_name, idx in col_map.items():
            val = row[idx]
            if val is None and col_name in ("currency", "legal_entity", "department", "country", "l1", "l2", "l3", "hibob_id", "display_name", "l2_function"):
                skip = True
                break
            cast = expected_columns[col_name]
            if val is not None and cast == bool:
                if isinstance(val, str):
                    val = val.lower() in ("yes", "true", "1", "y")
                else:
                    val = bool(val)
            elif val is not None and cast in (int, float):
                try:
                    val = cast(val)
                except (ValueError, TypeError):
                    val = None
            elif val is not None:
                val = str(val)
            data[col_name] = val
        if not skip:
            rows.append(data)

    if replace:
        db.query(model).delete()

    for row_data in rows:
        db.add(model(**row_data))

    db.commit()

    count = db.query(model).count()
    return {"table": table_name, "rows_imported": len(rows), "total_rows": count}
