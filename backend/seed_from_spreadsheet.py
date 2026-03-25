"""
One-time script to seed reference tables from the existing nShift spreadsheet.
Run from the backend directory:
    python seed_from_spreadsheet.py
"""
import sys
from pathlib import Path
from datetime import date

import openpyxl

# Add parent to path so we can import app modules
sys.path.insert(0, str(Path(__file__).parent))

from app.database import engine, SessionLocal, Base
from app.models.reference import (
    FxRate, LoadingMultiplier, DepartmentHierarchy,
    MeritAssumption, QuotaRampSchedule, LeaveCostRate,
)

SPREADSHEET = Path(__file__).parent.parent / "nShift Census Example anonymised.xlsx"


def seed():
    Base.metadata.create_all(bind=engine)
    db = SessionLocal()

    try:
        wb = openpyxl.load_workbook(str(SPREADSHEET), data_only=True)
        ws = wb["ref"]

        # --- FX Rates (col H=currency, I=rate) ---
        db.query(FxRate).delete()
        for row in ws.iter_rows(min_row=3, max_row=20, min_col=8, max_col=9, values_only=True):
            currency, rate = row
            if not currency or not isinstance(rate, (int, float)):
                continue
            if currency in ("xxx", "Tenur"):
                continue
            # Skip numeric "currencies" — these are quota ramp data in the same columns
            if isinstance(currency, (int, float)):
                continue
            db.add(FxRate(currency=str(currency), rate_to_eur=float(rate)))
        db.flush()
        print(f"FX Rates: {db.query(FxRate).count()} rows")

        # --- Loading Multipliers (col O=entity, P=location, Q=benefits, R=taxes, S=currency) ---
        # There are two sets separated by "effective until" row — take rows until first "effective until"
        db.query(LoadingMultiplier).delete()
        for row in ws.iter_rows(min_row=3, max_row=30, min_col=15, max_col=19, values_only=True):
            entity, location, benefits, taxes, currency = row
            if not entity or entity in ("xxx", "Legal Entity"):
                continue
            if entity == "effective until":
                break
            if not isinstance(benefits, (int, float)):
                continue
            db.add(LoadingMultiplier(
                legal_entity=str(entity),
                default_location=str(location) if location else None,
                employer_benefits_pct=float(benefits),
                employer_taxes_pct=float(taxes),
                currency=str(currency),
            ))
        db.flush()
        print(f"Loading Multipliers: {db.query(LoadingMultiplier).count()} rows")

        # --- Department Hierarchy (col B=dept, C=L1, D=L2, E=L3) ---
        db.query(DepartmentHierarchy).delete()
        seen_depts = set()
        for row in ws.iter_rows(min_row=3, max_row=200, min_col=2, max_col=5, values_only=True):
            dept, l1, l2, l3 = row
            if not dept or dept in ("(blank)", "xxx"):
                continue
            if not l1 or l1 == "xxx":
                continue
            dept_str = str(dept)
            if dept_str in seen_depts:
                continue
            seen_depts.add(dept_str)
            db.add(DepartmentHierarchy(
                department=dept_str,
                l1=str(l1),
                l2=str(l2),
                l3=str(l3),
            ))
        db.flush()
        print(f"Department Hierarchy: {db.query(DepartmentHierarchy).count()} rows")

        # --- Merit Assumptions (default) ---
        db.query(MeritAssumption).delete()
        db.add(MeritAssumption(
            merit_rate=0.03,
            effective_date=date(2026, 4, 1),
            joiner_cutoff_date=date(2025, 9, 29),
        ))
        print("Merit Assumptions: 1 row (default 3% on 1 Apr)")

        # --- Quota Ramp Schedule (from spreadsheet: 90d=30%, 180d=70%, 270d=100%) ---
        db.query(QuotaRampSchedule).delete()
        db.add(QuotaRampSchedule(day_from=0, day_to=89, ramp_pct=0.0))
        db.add(QuotaRampSchedule(day_from=90, day_to=179, ramp_pct=0.3))
        db.add(QuotaRampSchedule(day_from=180, day_to=269, ramp_pct=0.7))
        db.add(QuotaRampSchedule(day_from=270, day_to=None, ramp_pct=1.0))
        print("Quota Ramp Schedule: 4 rows")

        # --- Leave Cost Rates (placeholder defaults — user should update) ---
        db.query(LeaveCostRate).delete()
        countries = ["Sweden", "Norway", "Denmark", "UK", "Romania", "Netherlands",
                     "Belgium", "Spain", "Finland", "Poland", "Germany", "Austria", "Ireland"]
        for country in countries:
            db.add(LeaveCostRate(country=country, cost_pct=0.0))
        print(f"Leave Cost Rates: {len(countries)} rows (all 0% — update as needed)")

        db.commit()
        print("\nSeed complete.")

    finally:
        db.close()


if __name__ == "__main__":
    seed()
